from openpyxl import load_workbook
from itertools import zip_longest
from pathlib import Path
from spire.xls import *
from spire.xls.common import *


# excel functions
def unmerge_cells_keep_value(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        merged_cell_ranges = list(sheet.merged_cells.ranges)
        for merged_cell_range in merged_cell_ranges:
            min_row, min_col, max_row, max_col = (
                merged_cell_range.min_row,
                merged_cell_range.min_col,
                merged_cell_range.max_row,
                merged_cell_range.max_col,
            )
            merged_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(merged_cell_range))
            # Set the value to all cells in the merged range
            for row in sheet.iter_rows(
                min_row=merged_cell_range.min_row,
                max_row=merged_cell_range.max_row,
                min_col=merged_cell_range.min_col,
                max_col=merged_cell_range.max_col,
            ):
                for cell in row:
                    if cell.value is None:
                        cell.value = merged_cell_value
    wb.save(f"{Path(file_path).stem}_unmerge.xlsx")


def split_excel_sheets_into_files(excel_path):
    workbook = Workbook()
    workbook.LoadFromFile(excel_path)

    for worksheet in workbook.Worksheets:
        new_workbook = Workbook()

        # Clear the default worksheets in the new workbook
        new_workbook.Worksheets.Clear()

        # Copy the worksheets from the loaded Excel file to the new workbook
        new_workbook.Worksheets.AddCopy(worksheet)

        output_file_path = f"{worksheet.Name}.xlsx"
        new_workbook.SaveToFile(output_file_path, FileFormat.Version2016)
        print(f"***Split {output_file_path}****")

    workbook.Dispose()


def compare_excel_files(excel_path1, excel_path2):
    is_same = True
    wb1 = load_workbook(excel_path1)
    wb2 = load_workbook(excel_path2)

    sheet1 = wb1.active
    sheet2 = wb2.active

    max_rows = max(sheet1.max_row, sheet2.max_row)
    max_cols = max(sheet1.max_column, sheet2.max_column)

    for row_num in range(1, max_rows + 1):
        row_sheet1 = []
        row_sheet2 = []

        for col_num in range(1, max_cols + 1):
            cell_sheet1 = sheet1.cell(row=row_num, column=col_num)
            cell_sheet2 = sheet2.cell(row=row_num, column=col_num)

            row_sheet1.append(cell_sheet1.value)
            row_sheet2.append(cell_sheet2.value)

        for col_num, (c1, c2) in enumerate(
            zip_longest(row_sheet1, row_sheet2), start=1
        ):
            if c1 != c2:
                print(f"Row {row_num} Col {col_num} - {c1} != {c2}")
                if is_same:
                    is_same = False

        if sheet1.max_column < sheet2.max_column:
            print(f"Row {row_num} missing columns")
        elif sheet1.max_column > sheet2.max_column:
            print(f"Row {row_num} additional columns")
    if is_same:
        print(f"{excel_path1} && {excel_path2} have same datas")
    wb1.close()
    wb2.close()
